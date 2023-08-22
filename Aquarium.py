from PyQt5 import QtWidgets, uic
from PyQt5.QtSerialPort import QSerialPort, QSerialPortInfo
from PyQt5.QtCore import QIODevice
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import datetime


app = QtWidgets.QApplication([])
ui = uic.loadUi("AquariumUI.ui")
ui.setWindowTitle("SmartAquarium")
ui.tSlider.setVisible(False)
ui.phSlider.setVisible(False)
ui.tName.setVisible(False)
ui.phName.setVisible(False)

serial = QSerialPort()
serial.setBaudRate(115200)
portList = []
ports = QSerialPortInfo().availablePorts()
portList.append('Arduino Simulator')
for port in ports:
    portList.append(port.portName())
ui.cBox.addItems(portList)

wb = load_workbook('SmartAquarium.xlsx')
ws = wb.active
date=str(datetime.datetime.now().strftime("%d-%m-%y"+"_"+"%H-%M-%S"))
wscur = wb.create_sheet("Session " + date)


def tSliderChanged():
    a = ui.tSlider.value()
    if a<18:
        shlcd.display(int(100))
    else:
        shlcd.display(int(0))


def phSliderChanged():
    a = ui.phSlider.value()
    if a<5:
        clcd.display(int(100))
    else:
        clcd.display(int(0))


def dataBase():
    temp=ui.tSlider.value()
    ph=ui.phSlider.value()
    i=1
    global wb
    current_sheet = wb["Session " + date]

    while (ws['A'+str(i)]!=""):
        if (current_sheet.cell(row=i, column=1).value in [None,'None']):
            break
        i+=1

    current_sheet['A' + str(i)] = datetime.datetime.now()
    current_sheet['B' + str(i)].value = ui.cBox.currentText()
    current_sheet['C' + str(i)].value = temp
    current_sheet['D' + str(i)].value = ph
    if temp < 18:
        current_sheet['E' + str(i)].value = 'ON'
    else:
        current_sheet['E' + str(i)].value = 'OFF'
    if ph < 5:
        current_sheet['F' + str(i)].value = 'ON'
    else:
        current_sheet['F' + str(i)].value = 'OFF'
    wb.save("SmartAquarium.xlsx")


def onRead():
    if not serial.canReadLine(): return     # выходим если нечего читать
    rx = serial.readLine()
    rxs = str(rx, 'utf-8').strip()
    data = rxs.split(',')
    if data[0] != '':
        temp=int(((float(data[0]) * 5) / 1024.0) - 0.5) * 100
        tlcd.display(temp)
        if temp < 18:
            shlcd.display(int(100))
        else:
            shlcd.display(int(0))
    if data[1] != '':
        ph=int((float(data[1])*3.3/1024.0+1.1)*3.5)
        phlcd.display(ph)
        if ph < 5:
            clcd.display(int(100))
        else:
            clcd.display(int(0))


def onOpen():
    activePort = ui.cBox.currentText()
    if (activePort!='Arduino Simulator'):
        serial.setPortName(ui.cBox.currentText())
        serial.open(QIODevice.ReadWrite)
    else:
        ui.tSlider.setEnabled(True)
        ui.phSlider.setEnabled(True)
        ui.tSlider.setVisible(True)
        ui.phSlider.setVisible(True)
        ui.tName.setVisible(True)
        ui.phName.setVisible(True)
        ui.cBox.setEnabled(False)


def onClose():
    activePort = ui.cBox.currentText()
    if activePort != 'Arduino Simulator':
        serial.close()
    else:
        ui.tSlider.setEnabled(False)
        ui.phSlider.setEnabled(False)
        ui.cBox.setEnabled(True)


wscur['A1'] = "Time"
wscur['B1'] = "Port"
wscur['C1'] = "Temperature"
wscur['D1'] = "pH"
wscur['E1'] = "Healing System"
wscur['F1'] = "Compressor"

tlcd = ui.tLCD
phlcd = ui.phLCD
shlcd = ui.healingLCD
clcd = ui.compLCD

serial.readyRead.connect(onRead)
ui.openB.clicked.connect(onOpen)
ui.closeB.clicked.connect(onClose)

ui.tSlider.valueChanged.connect(tlcd.display)
ui.phSlider.valueChanged.connect(phlcd.display)
ui.tSlider.valueChanged.connect(tSliderChanged)
ui.phSlider.valueChanged.connect(phSliderChanged)
ui.tSlider.valueChanged.connect(dataBase)
ui.phSlider.valueChanged.connect(dataBase)



wb.save("SmartAquarium.xlsx")



ui.show()
app.exec()